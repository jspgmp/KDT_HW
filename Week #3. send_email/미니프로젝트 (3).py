'''
미니프로젝트 (3).ipynb 파일을 처음 보시는 분을 위해!

(1) 프로젝트 명 : 네이버 뉴스 자동 수집 후 메일 발송기

(2) 프로젝트 설명(200자 이내) : 특정 키워드에 해당하는 뉴스 기사를 네이버에서 찾아 기사 제목, 링크를 
엑셀 파일로 만들어 미리 만들어둔 엑셀 파일에 있는 메일링 대상자에게 메일로 전달합니다.

(3) 프로젝트 과제 상세 :

사용자가 원하는 키워드 입력받기
네이버 뉴스를 수집해 주는 모듈을 이용해서 해당 키워드 뉴스 수집 후 엑셀 파일에 제목, 링크, 요약문 기록하기
수집 데이터 엑셀 파일을 메일링 대상자 엑셀 파일을 읽어 대상자들에게 메일 보내기
(4) 점검 및 합격 기준표 :

해당 파트를 수강하면서 사용한 문법과 기능을 활용할 수 있어야 합니다.
전체적인 업무의 흐름을 이해하고 작성하여야 합니다.
각 단계마다 주어진 과제 코드를 완성하여야 합니다.


아래 코드를 실행해서 NaverNewsCrawler 모듈을 임포트하세요.

# 크롤러 코드를 위한 모듈 설치
!pip install requests
!pip install beautifulsoup4
'''

from NaverNewsCrawler import NaverNewsCrawler


####사용자로 부터 기사 수집을 원하는 키워드를 input을 이용해 입력받아 ? 부분에 넣으세요
choosekeyword = input('수집을 원하는 키워드를 입력하세요: ')
crawler = NaverNewsCrawler(choosekeyword)

#### 수집한 데이터를 저장할 엑셀 파일명을 input을 이용해 입력받아 ? 부분에 넣으세요
file_name = input('저장할 파일명을 입력하세요: ')
if '.' in file_name:
    file_name = file_name[:file_name.index('.')] ##파일명에 확장자를 붙일경우 잘라내고 다시새로붙임
crawler.get_news(file_name + '.xlsx')

#### 아래코드를 실행해 이메일 발송 기능에 필요한 모듈을 임포트하세요.
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import re
import json ##json 사용을 위해 import

with open('conf.json')as f: ##json 으로 정보를 빼기 위한 장치
    confi = json.load(f)

#### gmail 발송 기능에 필요한 계정 정보를 아래 코드에 입력하세요.
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
SMTP_USER = confi['email']## json파일로 이메일과 비밀번호 뺴기
SMTP_PASSWORD = confi['password']

#### 아래 코드를 실행해 메일 발송에 필요한 send_mail 함수를 만드세요.
def send_mail(name, addr, subject, contents, attachment=None):
    if not re.match('(^[a-zA-Z0-9_.-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        print('Wrong email')
        return

    msg = MIMEMultipart('alternative')
    if attachment:
        msg = MIMEMultipart('mixed')

    msg['From'] = SMTP_USER
    msg['To'] = addr
    msg['Subject'] = name + '님, \n' + subject + '에 대해 알아보세요'
    text = MIMEText(contents, _charset='utf-8')
    msg.attach(text)

    if attachment:
        from email.mime.base import MIMEBase
        from email import encoders

        file_data = MIMEBase('application', 'octet-stream')
        file_data.set_payload(open(attachment, 'rb').read())
        encoders.encode_base64(file_data)

        import os
        filename = os.path.basename(attachment)
        file_data.add_header('Content-Disposition', 'attachment', filename=filename)##한글파일도 정상첨부 하기위해 수정
        msg.attach(file_data)

    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    smtp.close()

#### 프로젝트 폴더에 있는 email_list.xlsx 파일에 이메일 받을 사람들의 정보를 입력하세요.
from openpyxl import Workbook

emailList_in = []
emailList_in.append(['미스터A', 'mmriann@gmail.com'])
emailList_in.append(['미스터B', 'mmriann@Naver.com'])

wb_add = Workbook()
ws_add = wb_add.active
ws_add['A1'] = "email_list"
ws_add.append(['번호', '이름', '이메일'])

numstart = 1
for email_in in emailList_in :
    ws_add.append([numstart, email_in[0], email_in[1]])
    numstart += 1
wb_add.save('email_list.xlsx')

#### 엑셀 파일의 정보를 읽어올 수 있는 모듈을 import하세요.
from openpyxl import load_workbook

#### email_list.xlsx 파일을 읽어와 해당 사람들에게 수집한 뉴스 정보 엑셀 파일을 send_mail 함수를 이용해 전송하세요.
wb_loadEmail = load_workbook('email_list.xlsx', read_only=True) 
ws_loadEmail = wb_loadEmail.active

for row in ws_loadEmail.iter_rows(min_row = 3, max_col = 3) :
    receiver_name = row[1].value
    receiver_email = row[2].value
    contents = (f'안녕하세요? {receiver_name}님 \n{choosekeyword}에 대한 정보를 보내드립니다.')
    send_mail(receiver_name, receiver_email, choosekeyword, contents, f'{file_name}.xlsx')
    print(f'{receiver_name} 님에게 {choosekeyword}에 대한 내용을 전송하였습니다.')
